# -*- coding: utf-8 -*- 

import wx
import os
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference

class MyFrame1 ( wx.Frame ):
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = wx.EmptyString, pos = wx.DefaultPosition, size = wx.Size( 842,416 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
        
        self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )
        
        bSizer1 = wx.BoxSizer( wx.VERTICAL )
        
        self.m_staticText1 = wx.StaticText( self, wx.ID_ANY, u"출생 데이터 분석", wx.DefaultPosition, wx.DefaultSize, wx.ALIGN_CENTRE )
        self.m_staticText1.Wrap( -1 )
        self.m_staticText1.SetFont( wx.Font( 20, 70, 90, 90, False, wx.EmptyString ) )
        
        bSizer1.Add( self.m_staticText1, 0, wx.ALL|wx.EXPAND, 5 )
        
        bSizer2 = wx.BoxSizer( wx.HORIZONTAL )
        
        bSizer2.SetMinSize( wx.Size( -1,20 ) ) 
        self.m_staticText2 = wx.StaticText( self, wx.ID_ANY, u"Data Folder :", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText2.Wrap( -1 )
        bSizer2.Add( self.m_staticText2, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
        
        self.m_dirPicker1 = wx.DirPickerCtrl( self, wx.ID_ANY, wx.EmptyString, u"Select a folder", wx.DefaultPosition, wx.DefaultSize, wx.DIRP_CHANGE_DIR|wx.DIRP_DEFAULT_STYLE|wx.DIRP_DIR_MUST_EXIST )
        self.m_dirPicker1.SetMinSize( wx.Size( 350,-1 ) )
        
        bSizer2.Add( self.m_dirPicker1, 1, wx.ALL, 5 )
        
        self.go = wx.Button( self, wx.ID_ANY, u"분석", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer2.Add( self.go, 0, wx.ALL, 5 )
        
        
        bSizer1.Add( bSizer2, 0, wx.EXPAND, 5 )
        
        self.console = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,120 ), wx.HSCROLL|wx.TE_MULTILINE )
        bSizer1.Add( self.console, 1, wx.ALL|wx.EXPAND, 5 )
        
        
        self.SetSizer( bSizer1 )
        self.Layout()
        
        self.Centre( wx.BOTH )
        
        # Connect Events
        self.go.Bind( wx.EVT_BUTTON, self.loading )
    
    def __del__( self ):
        pass
    
    # Virtual event handlers, overide them in your derived class
    def loading( self, event ):
        self.console.SetValue('')
        global thisFolder
        global thisFile
        years = []
        w = []
        thisFolder = self.m_dirPicker1.GetPath()
        wholeFile = os.listdir(thisFolder)
        global csvs 
        csvs = [file for file in wholeFile if file.endswith(".csv")]
        csvnum=0
        s_data=[]
        for csv in csvs:
            thisFile = csv
            yr = thisFile[:4]
            years.append(yr)
            theFile=thisFolder+"\\"+thisFile
            with open(theFile,'r',encoding='cp949') as f:
                thisLine = f.readlines()
                zippedData=[]
                colNames = []
                rowN = 0
                for rowData in thisLine:
                    x = rowData.replace('","','"|"')
                    x = x.replace('"','')
                    if x[-1] == '\n': x = x[:-1] 
                    if rowN==0:
                        cols = x.split('|')
                        for y in cols:
                            if "_" not in y :
                                colNames.append(y)
                            else:
                                z = y.find("_")
                                colName=y[z-3:z+2]
                                colNames.append(colName)
                                
                    else:
                        indexx = x.find('(')
                        rowName = x[:indexx-2]
                        colData = x.split('|')
                        colData[0] = rowName
                        zippedData.append(zip(colNames,colData))
                    rowN += 1
                    
                birthm = []
                birthf = []    
                births = []    
                yrData = []
                dataNum = 0
                for i in zippedData:
                    birthm.append(0)
                    birthf.append(0)
                    births.append(0)
                    for j, zd in enumerate(i):
                        x = zd[1].replace(',', '')
                        if zd[0].endswith('남') : birthm[dataNum]+=int(x)
                        if zd[0].endswith('여') : birthf[dataNum]+=int(x)
                        if zd[0].endswith('계') : births[dataNum]+=int(x)
                        if zd[0]=='행정구역' : g=x
                        w.append(j)
                    datum = (g,birthm[dataNum],birthf[dataNum],births[dataNum])
                    if g == "전국" : lastRow = datum 
                    else : yrData.append(datum)
                    dataNum+=1
                yrData.append(lastRow)
                y_data = pd.DataFrame(yrData)
                y_data[1] = y_data[1].astype(int)
                y_data[2] = y_data[2].astype(int)
                y_data[3] = y_data[3].astype(int)
                s_data.append(y_data)
                csvnum+=1
            self.console.AppendText("분석 중 : "+thisFile+"\n")
        excelResult = '%s~%s.xlsx' %(years[0],years[-1])
        with pd.ExcelWriter(excelResult) as writer:
            for i in range(csvnum):
                s_data[i].to_excel(writer, sheet_name = years[i], header=[years[i]+"년",'남','여','계'],startcol=-1)
            self.console.AppendText(years[0]+"년 부터 "+years[-1]+"년 까지의 통계가 생성되었습니다.\n")

        res = openpyxl.load_workbook(excelResult, data_only=True)
        sheet_list = res.sheetnames
        sheet = []
        for i in range(len(sheet_list)):
            sheet.append(res[sheet_list[i]])
            chart = BarChart()
            chart.title = years[i]+"년 시도별 출생신고 아동 수"
            datas = Reference(sheet[i], min_col=2, max_col=4, min_row=1, max_row=sheet[i].max_row-1)
            categ = Reference(sheet[i], min_col=1, max_col=1, min_row=2, max_row=sheet[i].max_row-1)
            chart.add_data(datas, titles_from_data=True)
            chart.set_categories(categ)
            chart.x_axis.title = "행정구역"
            chart.y_axis.title = "출생신고수 (명)"
            chart.style = 2
            sheet[i].add_chart(chart, "E2")
            chart.height = 10
            chart.width = 17
        res.save(excelResult)
        
        
###############################################################################
if __name__ == '__main__':
    ex = wx.App(False)
    frame = MyFrame1(None)
    frame.Show(True)
    ex.MainLoop()
